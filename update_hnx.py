"""
update_hnx.py
=============
Incremental update script for the HNX bond auction results baseline file.

Reads the most recent auction date already present in ket_qua_dau_thau_nen.xlsx,
fetches only the new records published on HNX since that date, merges them
in (deduplicated), and re-saves the formatted Excel file.

Run:
    python update_hnx.py

Requires ket_qua_dau_thau_nen.xlsx to already exist in the same folder
(built once via scrape_hnx.py).
"""

import os
os.environ["PYTHONIOENCODING"] = "utf-8"  # avoid encoding errors when printing Vietnamese text

from playwright.sync_api import sync_playwright
import pandas as pd
import re
import io
from datetime import datetime
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

URL = "https://hnx.vn/trai-phieu/ket-qua-dau-thau.html"
OUTPUT_FILE = "ket_qua_dau_thau_nen.xlsx"


def get_last_date_in_file():
    """Read the existing baseline file and return (dataframe, latest auction date)."""
    df = pd.read_excel(OUTPUT_FILE, engine="openpyxl")
    df["Ngày TCPH"] = pd.to_datetime(df["Ngày TCPH"])
    return df, df["Ngày TCPH"].max()


def scrape_range(from_date, to_date):
    """Fetch all auction results published between from_date and to_date
    (both in dd/mm/yyyy format)."""
    all_data = []

    with sync_playwright() as p:
        # channel="msedge" uses the Microsoft Edge browser already installed
        # on the machine, avoiding the need to download a separate Chromium
        # build (useful on corporate machines that block new browser installs).
        browser = p.chromium.launch(channel="msedge", headless=False)
        page = browser.new_page()
        page.goto(URL, timeout=60000)
        page.wait_for_timeout(2000)

        page.evaluate("next_to_ThongBaoKetQuaDauThau()")
        page.wait_for_function(
            "document.querySelector('#divDataTables') && "
            "document.querySelector('#divDataTables').innerText.includes('Đợt đấu thầu')",
            timeout=15000
        )
        print("Entered correct tab.")

        page.evaluate(f"""
            document.querySelector('#txtFromDate').value = '{from_date}';
            document.querySelector('#txtToDate').value = '{to_date}';
        """)
        page.evaluate("TableSeach(1, 1)")
        page.wait_for_timeout(2500)

        full_text = page.locator("#divDataTables").inner_text()
        match = re.search(r"(\d[\d\.]*)\s*bản ghi", full_text)
        total_records = int(match.group(1).replace(".", "")) if match else None
        print(f"Total new records: {total_records}")

        if not total_records:
            browser.close()
            return []

        page_num = 1
        while True:
            if page_num > 1:
                page.evaluate(f"TableSeach({page_num}, 0)")
                page.wait_for_timeout(1800)

            html_fragment = page.locator("#_tableDatas").evaluate("el => el.outerHTML")

            try:
                df = pd.read_html(io.StringIO(html_fragment), match="Tên TCPH", thousands=None)[0]
                all_data.append(df)
                print(f"Page {page_num}: {len(df)} rows")
            except ValueError as e:
                print(f"Page {page_num}: read error ({e}), stopping.")
                break

            total_collected = sum(len(d) for d in all_data)
            if total_collected >= total_records:
                print(f"Collected {total_collected}/{total_records}, done.")
                break

            page_num += 1
            if page_num > 300:
                print("Exceeded 300 pages, stopping as a safety limit.")
                break

        browser.close()

    return all_data


def convert_vn_number(series):
    """Convert Vietnamese-formatted numbers (dot = thousands separator,
    comma = decimal separator) into proper floats. Any unparseable value
    (e.g. '-', blank) becomes NaN instead of raising an error."""
    cleaned = (
        series.astype(str)
        .str.strip()
        .replace({"": None, "nan": None, "None": None, "-": None})
        .str.replace(".", "", regex=False)
        .str.replace(",", ".", regex=False)
    )
    return pd.to_numeric(cleaned, errors="coerce")


def clean_new_data(new_df):
    """Apply the same cleaning/typing rules used when building the baseline file."""
    numeric_cols = [
        "GT gọi thầu", "GT đặt thầu", "GT trúng thầu",
        "Số tiền thanh toán trái phiếu trúng thầu",
        "GT gọi thầu phát hành thêm", "GT đặt thầu phát hành thêm",
        "GT trúng thầu phát hành thêm", "Số tiền thanh toán trái phiếu phát hành thêm",
        "Lãi suất danh nghĩa (%/Năm)", "Lãi suất trúng thầu (%/Năm)",
        "LS đăng ký thấp nhất (%/Năm)", "LS đăng ký cao nhất (%/Năm)"
    ]
    for col in numeric_cols:
        if col in new_df.columns:
            new_df[col] = convert_vn_number(new_df[col])

    if "Kỳ hạn" in new_df.columns:
        new_df["Bond Term (Years)"] = new_df["Kỳ hạn"].astype(str).str.extract(r"(\d+)").astype(float)
        new_df["Bond Term (Months)"] = new_df["Bond Term (Years)"] * 12

    date_col = "Ngày TCPH ▼" if "Ngày TCPH ▼" in new_df.columns else "Ngày TCPH"
    if date_col in new_df.columns:
        new_df["Ngày TCPH"] = pd.to_datetime(
            new_df[date_col].astype(str).str.replace(" ▼", "", regex=False),
            format="%d/%m/%Y", errors="coerce"
        )
        if date_col != "Ngày TCPH":
            new_df.drop(columns=[date_col], inplace=True)

    if "Ngày phát hành" in new_df.columns:
        new_df["Ngày phát hành"] = pd.to_datetime(new_df["Ngày phát hành"], format="%d/%m/%Y", errors="coerce")

    return new_df


def save_formatted_excel(full_df, output_file):
    """Write the DataFrame to Excel with number formatting, frozen header
    row, and a native Excel Table (with filters) applied."""
    full_df.sort_values("Ngày TCPH", inplace=True)
    full_df.reset_index(drop=True, inplace=True)

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        full_df.to_excel(writer, index=False, sheet_name="Data")
        ws = writer.sheets["Data"]
        ws.freeze_panes = "A2"

        money_cols = [
            "GT gọi thầu", "GT đặt thầu", "GT trúng thầu",
            "Số tiền thanh toán trái phiếu trúng thầu",
            "GT gọi thầu phát hành thêm", "GT đặt thầu phát hành thêm",
            "GT trúng thầu phát hành thêm", "Số tiền thanh toán trái phiếu phát hành thêm",
        ]
        rate_cols = [
            "Lãi suất danh nghĩa (%/Năm)", "Lãi suất trúng thầu (%/Năm)",
            "LS đăng ký thấp nhất (%/Năm)", "LS đăng ký cao nhất (%/Năm)"
        ]
        date_cols = ["Ngày TCPH", "Ngày phát hành"]

        for idx, col_name in enumerate(full_df.columns, start=1):
            col_letter = get_column_letter(idx)
            ws.column_dimensions[col_letter].width = max(14, len(str(col_name)) + 2)
            if col_name in money_cols:
                for cell in ws[col_letter][1:]:
                    cell.number_format = "#,##0"
            elif col_name in rate_cols:
                for cell in ws[col_letter][1:]:
                    cell.number_format = "0.00"
            elif col_name in date_cols:
                for cell in ws[col_letter][1:]:
                    cell.number_format = "dd/mm/yyyy"

        n_rows = full_df.shape[0] + 1
        n_cols = full_df.shape[1]
        last_col_letter = get_column_letter(n_cols)
        table = Table(displayName="KetQuaDauThau", ref=f"A1:{last_col_letter}{n_rows}")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
        ws.add_table(table)


def main():
    if not os.path.exists(OUTPUT_FILE):
        print(f"{OUTPUT_FILE} not found. Copy the baseline file into this folder first, "
              f"or run scrape_hnx.py to build it from scratch.")
        return

    old_df, last_date = get_last_date_in_file()
    print(f"Existing data goes up to: {last_date.strftime('%d/%m/%Y')}")

    from_date = last_date.strftime("%d/%m/%Y")
    to_date = datetime.now().strftime("%d/%m/%Y")

    new_data = scrape_range(from_date, to_date)
    if not new_data:
        print("No new data to update.")
        return

    new_df = pd.concat(new_data, ignore_index=True)
    new_df = clean_new_data(new_df)

    combined = pd.concat([old_df, new_df], ignore_index=True)
    # Deduplicate on full-row match (safer than a subset of columns, since
    # some legitimate records share the same auction code but differ in
    # other fields, e.g. supplementary issuances).
    combined.drop_duplicates(inplace=True)

    save_formatted_excel(combined, OUTPUT_FILE)
    added = len(combined) - len(old_df)
    print(f"DONE: +{added} new rows, {len(combined)} total, saved to {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
