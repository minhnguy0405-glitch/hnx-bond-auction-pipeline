"""
scrape_hnx.py
=============
One-time script to build the baseline Excel file of Vietnamese government
bond auction results from the Hanoi Stock Exchange (HNX), covering
01/01/2016 to present.

Run:
    python scrape_hnx.py

Output:
    ket_qua_dau_thau_nen.xlsx  (~2,400+ rows, formatted and ready to use)

This script only needs to be run ONCE to create the baseline file.
For ongoing updates, use update_hnx.py instead (much faster, only
fetches new records since the last run).
"""

import os
os.environ["PYTHONIOENCODING"] = "utf-8"  # avoid encoding errors when printing Vietnamese text

from playwright.sync_api import sync_playwright
import pandas as pd
import re
import io
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

URL = "https://hnx.vn/trai-phieu/ket-qua-dau-thau.html"
OUTPUT_FILE = "ket_qua_dau_thau_nen.xlsx"
FROM_DATE = "01/01/2016"
TO_DATE = "14/07/2026"  # update to current date when re-running a full rebuild


def scrape_all_pages():
    """Navigate the HNX bond auction results page and collect every page
    of results within the configured date range."""
    all_data = []

    with sync_playwright() as p:
        # channel="msedge" uses the Microsoft Edge browser already installed
        # on the machine, avoiding the need to download a separate Chromium
        # build (useful on corporate machines that block new browser installs).
        browser = p.chromium.launch(channel="msedge", headless=False)
        page = browser.new_page()
        page.goto(URL, timeout=60000)
        page.wait_for_timeout(2000)

        # The page loads the "Bidding Calendar" tab by default; switch to the
        # "Bidding Results" tab using the site's own JS function.
        page.evaluate("next_to_ThongBaoKetQuaDauThau()")
        page.wait_for_function(
            "document.querySelector('#divDataTables') && "
            "document.querySelector('#divDataTables').innerText.includes('Đợt đấu thầu')",
            timeout=15000
        )
        print("Entered correct tab.")

        # Set the date range directly via JS (page.fill() times out because
        # the input is not always visible at this point in the page's
        # transition animation).
        page.evaluate(f"""
            document.querySelector('#txtFromDate').value = '{FROM_DATE}';
            document.querySelector('#txtToDate').value = '{TO_DATE}';
        """)
        page.evaluate("TableSeach(1, 1)")
        page.wait_for_timeout(2500)

        full_text = page.locator("#divDataTables").inner_text()
        match = re.search(r"(\d[\d\.]*)\s*bản ghi", full_text)
        total_records = int(match.group(1).replace(".", "")) if match else None
        print(f"Total records: {total_records}")

        page_num = 1
        while True:
            if page_num > 1:
                page.evaluate(f"TableSeach({page_num}, 0)")
                page.wait_for_timeout(1800)

            html_fragment = page.locator("#_tableDatas").evaluate("el => el.outerHTML")

            try:
                # thousands=None is important: by default pandas treats ","
                # as a thousands separator, which corrupts Vietnamese-style
                # decimal numbers (e.g. "4,18" would otherwise be misread).
                df = pd.read_html(io.StringIO(html_fragment), match="Tên TCPH", thousands=None)[0]
                all_data.append(df)
                print(f"Page {page_num}: {len(df)} rows")
            except ValueError as e:
                print(f"Page {page_num}: read error ({e}), stopping.")
                break

            total_collected = sum(len(d) for d in all_data)
            if total_records and total_collected >= total_records:
                print(f"Collected {total_collected}/{total_records}, done.")
                break

            page_num += 1
            if page_num > 300:
                print("Exceeded 300 pages, stopping as a safety limit.")
                break

        browser.close()

    return all_data


def convert_vn_number(series):
    """Convert Vietnamese-formatted numbers (dot = thousands separator,
    comma = decimal separator) into proper floats. Any unparseable value
    (e.g. '-', blank) becomes NaN instead of raising an error."""
    cleaned = (
        series.astype(str)
        .str.strip()
        .replace({"": None, "nan": None, "None": None, "-": None})
        .str.replace(".", "", regex=False)
        .str.replace(",", ".", regex=False)
    )
    return pd.to_numeric(cleaned, errors="coerce")


def clean_dataframe(full_df):
    """Deduplicate, convert numeric/date columns, and sort chronologically."""
    full_df.drop_duplicates(inplace=True)

    numeric_cols = [
        "GT gọi thầu", "GT đặt thầu", "GT trúng thầu",
        "Số tiền thanh toán trái phiếu trúng thầu",
        "GT gọi thầu phát hành thêm", "GT đặt thầu phát hành thêm",
        "GT trúng thầu phát hành thêm", "Số tiền thanh toán trái phiếu phát hành thêm",
        "Lãi suất danh nghĩa (%/Năm)", "Lãi suất trúng thầu (%/Năm)",
        "LS đăng ký thấp nhất (%/Năm)", "LS đăng ký cao nhất (%/Năm)"
    ]
    for col in numeric_cols:
        if col in full_df.columns:
            full_df[col] = convert_vn_number(full_df[col])

    # Derive bond term in years/months from the "Kỳ hạn" text column (e.g. "5 Năm")
    if "Kỳ hạn" in full_df.columns:
        full_df["Bond Term (Years)"] = full_df["Kỳ hạn"].astype(str).str.extract(r"(\d+)").astype(float)
        full_df["Bond Term (Months)"] = full_df["Bond Term (Years)"] * 12

    # Convert the auction date column to a proper datetime (source column
    # name may carry a sort-arrow character "▼" which needs stripping first)
    date_col = "Ngày TCPH ▼" if "Ngày TCPH ▼" in full_df.columns else "Ngày TCPH"
    if date_col in full_df.columns:
        full_df["Ngày TCPH"] = pd.to_datetime(
            full_df[date_col].astype(str).str.replace(" ▼", "", regex=False),
            format="%d/%m/%Y", errors="coerce"
        )
        if date_col != "Ngày TCPH":
            full_df.drop(columns=[date_col], inplace=True)

    if "Ngày phát hành" in full_df.columns:
        full_df["Ngày phát hành"] = pd.to_datetime(full_df["Ngày phát hành"], format="%d/%m/%Y", errors="coerce")

    full_df.sort_values("Ngày TCPH", inplace=True)
    full_df.reset_index(drop=True, inplace=True)
    return full_df


def save_formatted_excel(full_df, output_file):
    """Write the DataFrame to Excel with number formatting, frozen header
    row, and a native Excel Table (with filters) applied."""
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        full_df.to_excel(writer, index=False, sheet_name="Data")
        ws = writer.sheets["Data"]
        ws.freeze_panes = "A2"

        money_cols = [
            "GT gọi thầu", "GT đặt thầu", "GT trúng thầu",
            "Số tiền thanh toán trái phiếu trúng thầu",
            "GT gọi thầu phát hành thêm", "GT đặt thầu phát hành thêm",
            "GT trúng thầu phát hành thêm", "Số tiền thanh toán trái phiếu phát hành thêm",
        ]
        rate_cols = [
            "Lãi suất danh nghĩa (%/Năm)", "Lãi suất trúng thầu (%/Năm)",
            "LS đăng ký thấp nhất (%/Năm)", "LS đăng ký cao nhất (%/Năm)"
        ]
        date_cols = ["Ngày TCPH", "Ngày phát hành"]

        for idx, col_name in enumerate(full_df.columns, start=1):
            col_letter = get_column_letter(idx)
            ws.column_dimensions[col_letter].width = max(14, len(str(col_name)) + 2)
            if col_name in money_cols:
                for cell in ws[col_letter][1:]:
                    cell.number_format = "#,##0"
            elif col_name in rate_cols:
                for cell in ws[col_letter][1:]:
                    cell.number_format = "0.00"
            elif col_name in date_cols:
                for cell in ws[col_letter][1:]:
                    cell.number_format = "dd/mm/yyyy"

        n_rows = full_df.shape[0] + 1
        n_cols = full_df.shape[1]
        last_col_letter = get_column_letter(n_cols)
        table = Table(displayName="KetQuaDauThau", ref=f"A1:{last_col_letter}{n_rows}")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
        ws.add_table(table)


def main():
    all_data = scrape_all_pages()
    if not all_data:
        print("NO DATA collected.")
        return

    full_df = pd.concat(all_data, ignore_index=True)
    full_df = clean_dataframe(full_df)
    save_formatted_excel(full_df, OUTPUT_FILE)

    print(f"DONE: {len(full_df)} rows saved to {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
